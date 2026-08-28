# -*- coding: utf-8 -*-
"""Capa de datos: carga, cacheo y transformación del catálogo Glocalminds."""
import re
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

EXCEL_PATH = (
    Path(__file__).resolve().parent.parent
    / "data"
    / "experiencia_glocal_catálogo_web_histórico_hasta_10_Agosto_2026.xlsx"
)

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}
_DATE_RE = re.compile(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", re.IGNORECASE)

# Variantes detectadas del mismo método, escritas distinto por distintos lotes de codificación.
_METODOLOGIA_ALIASES = {
    "otra: café pro-acción": "Otra: Café ProAcción",
    "otra: café proacción": "Otra: Café ProAcción",
    "otra: café pro acción": "Otra: Café ProAcción",
    "otra: diseño de acción sabia": "Otra: Diseño para la Acción Sabia",
    "otra: tres horizontes": "Otra: Marco de los Tres Horizontes",
}

MULTILABEL_COLS = [
    "categorias", "categoria_macro", "metodologia", "actores", "actores_normalizados",
    "eje_gcaa", "objetivo_gcaa",
    "atributos_resiliencia", "subatributos_resiliencia",
    "beneficiarios_directos", "beneficiarios_indirectos",
]

NON_VALUES = {"no aplica", "no especificado", ""}

# Taxonomía oficial: Race to Resilience Technical Secretariat (julio 2023),
# "Introduction to Resilience Attributes, Their Subcategories, and Their Role
# in the Race to Resilience Campaign" (fuente CR2). 7 atributos, 19 sub-atributos.
RESILIENCE_TAXONOMY = {
    "1. Preparación y planificación": ["S1.1 - Preparación", "S1.2 - Planificación"],
    "2. Aprendizaje": ["S2.1 - Aprendizaje experiencial", "S2.2 - Aprendizaje educativo"],
    "3. Agencia": ["S3.1 - Autonomía", "S3.2 - Liderazgo", "S3.3 - Toma de decisiones"],
    "4. Colaboración social": ["S4.1 - Participación colectiva", "S4.2 - Conectividad", "S4.3 - Coordinación"],
    "5. Flexibilidad": ["S5.1 - Diversidad", "S5.2 - Redundancia"],
    "6. Equidad": ["S6.1 - Equidad distributiva", "S6.2 - Equidad de acceso"],
    "7. Activos": [
        "S7.1 - Finanzas", "S7.2 - Infraestructura", "S7.3 - Recursos naturales",
        "S7.4 - Tecnologías", "S7.5 - Servicios básicos",
    ],
}


def parse_spanish_date(value):
    if not isinstance(value, str) or not value.strip():
        return pd.NaT
    m = _DATE_RE.search(value)
    if not m:
        return pd.NaT
    day, month_name, year = m.groups()
    month = MESES.get(month_name.lower())
    if not month:
        return pd.NaT
    try:
        return pd.Timestamp(year=int(year), month=month, day=int(day))
    except ValueError:
        return pd.NaT


def _normalize_label(label):
    label = label.strip()
    return _METODOLOGIA_ALIASES.get(label.lower(), label)


def _normalize_multilabel_text(value):
    if not isinstance(value, str) or not value.strip():
        return value
    parts, seen = [], set()
    for p in value.split(";"):
        p = _normalize_label(p)
        if p and p not in seen:
            seen.add(p)
            parts.append(p)
    return "; ".join(parts)


@st.cache_data(show_spinner="Cargando catálogo de experiencias...")
def load_noticias() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_PATH, sheet_name="Noticias", engine="openpyxl")
    df = df.dropna(subset=["titulo"]).reset_index(drop=True)
    df.insert(0, "item", range(1, len(df) + 1))

    df["metodologia"] = df["metodologia"].apply(_normalize_multilabel_text)

    if "fecha_publicacion_web" in df.columns:
        df["fecha_parsed"] = pd.to_datetime(df["fecha_publicacion_web"], errors="coerce")
    else:
        df["fecha_parsed"] = pd.NaT
    if df["fecha_parsed"].isna().any():
        fecha_raw = df["fecha_publicacion"].fillna(df.get("fecha_catalogo"))
        fallback = fecha_raw.apply(parse_spanish_date)
        df["fecha_parsed"] = df["fecha_parsed"].fillna(fallback)
    df["anio"] = df["fecha_parsed"].dt.year
    df["tiene_fecha"] = df["fecha_parsed"].notna()

    for col in MULTILABEL_COLS:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("")

    return df


@st.cache_data(show_spinner=False)
def load_mapa_ubicaciones() -> pd.DataFrame:
    return pd.read_excel(EXCEL_PATH, sheet_name="Mapa_Ubicaciones", engine="openpyxl")


@st.cache_data(show_spinner=False)
def load_cuencas() -> pd.DataFrame:
    return pd.read_excel(EXCEL_PATH, sheet_name="Cuencas", engine="openpyxl")


@st.cache_data(show_spinner=False)
def load_subcuencas() -> pd.DataFrame:
    return pd.read_excel(EXCEL_PATH, sheet_name="Subcuencas", engine="openpyxl")


@st.cache_data(show_spinner=False)
def load_subsubcuencas() -> pd.DataFrame:
    return pd.read_excel(EXCEL_PATH, sheet_name="Subsubcuencas", engine="openpyxl")


@st.cache_data(show_spinner=False)
def load_sheet_as_text(sheet_name: str) -> list[str]:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[sheet_name]
    return [(row[0] or "") for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]


def explode_multilabel(df: pd.DataFrame, col: str, extra_cols=None) -> pd.DataFrame:
    """Una fila por cada (item, etiqueta) de una columna separada por ';'."""
    extra_cols = extra_cols or []
    sub = df[["item", col] + extra_cols].copy()
    sub[col] = sub[col].fillna("").astype(str).apply(
        lambda v: [p.strip() for p in v.split(";") if p.strip()]
    )
    sub = sub.explode(col)
    sub = sub[sub[col].notna() & (sub[col].astype(str) != "")]
    return sub.rename(columns={col: "label"})


def get_options(df: pd.DataFrame, col: str, exclude_non_values: bool = True) -> list[str]:
    exploded = explode_multilabel(df, col)
    vals = exploded["label"].dropna().unique().tolist()
    if exclude_non_values:
        vals = [v for v in vals if v.strip().lower() not in NON_VALUES]
    return sorted(vals)


def filter_by_multilabel(df: pd.DataFrame, col: str, selected: list[str]) -> pd.DataFrame:
    if not selected:
        return df
    selected_set = set(selected)

    def _match(v):
        parts = {p.strip() for p in str(v).split(";") if p.strip()}
        return bool(parts & selected_set)

    return df[df[col].apply(_match)]


def resilience_counts(df: pd.DataFrame):
    """Conteos de atributos y sub-atributos de resiliencia, incluyendo 0 para los que no aparecen en el catálogo."""
    a_exp = explode_multilabel(df, "atributos_resiliencia")
    a_exp = a_exp[a_exp["label"].str.lower() != "no aplica"]
    a_counts = a_exp["label"].value_counts()

    s_exp = explode_multilabel(df, "subatributos_resiliencia")
    s_exp = s_exp[s_exp["label"].str.lower() != "no aplica"]
    s_counts = s_exp["label"].value_counts()

    df_attr = pd.DataFrame(
        [{"atributo": attr, "n": int(a_counts.get(attr, 0))} for attr in RESILIENCE_TAXONOMY]
    )
    df_sub = pd.DataFrame(
        [
            {"atributo": attr, "subatributo": sub, "n": int(s_counts.get(sub, 0))}
            for attr, subs in RESILIENCE_TAXONOMY.items()
            for sub in subs
        ]
    )
    return df_attr, df_sub


def extract_country(display_name: str) -> str | None:
    """Aproxima el país a partir del último segmento de una dirección OSM."""
    if not isinstance(display_name, str) or not display_name.strip():
        return None
    last = display_name.split(",")[-1].strip()
    return last or None


COLUMN_LABELS = {
    "categorias": "Categoría temática",
    "categoria_macro": "Categoría macro",
    "metodologia": "Metodología de facilitación",
    "actores": "Actores institucionales",
    "actores_normalizados": "Actores institucionales (normalizado)",
    "eje_gcaa": "Eje GCAA",
    "objetivo_gcaa": "Objetivo GCAA",
    "atributos_resiliencia": "Atributo de resiliencia",
    "subatributos_resiliencia": "Sub-atributo de resiliencia",
    "beneficiarios_directos": "Beneficiarios directos",
    "beneficiarios_indirectos": "Beneficiarios indirectos",
    "enfoque_genero": "Enfoque de género",
    "lugar": "Lugar",
}
